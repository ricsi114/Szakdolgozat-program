# importing openpyxl and random
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import random


# Getting all the ingredient names from the first row to the "names" list.
wb = load_workbook('DatasetNew.xlsx')
ws = wb.active
names = []

for col in range(2, ws.max_column+1):
    names.append(ws[get_column_letter(col) + "1"].value)

# Creating the "Name".xlsx
wb = Workbook()
ws = wb.active
ws.append(names)


Beer_Count = int(input("How many beers do you want to generate?: "))
File_Name = input("Enter the Name of the File: ")
for row in range(2, Beer_Count+2):
    for col in range(1, len(names)+1):
        ws[get_column_letter(col) + str(row)].value = 0

# Setting up the variables for the randomizers

Base_Mals = ["Extra Pale", "Maris Otter", "Extra Pale Maris Otter"]
Yeasts = ["Wyeast 1056 - American Aleâ„¢",	"Wyeast 1272 - American Ale IIâ„¢",	"Wyeast American Ale II Strain 1272", "US-05", "Wyeast 3522 - Ardennesâ„¢", "Wyeast 3944 - Belgian Witbierâ„¢", "Wyeast 3711 - French Saisonâ„¢", "Wyeast 1388 â€“ Belgian Strong Aleâ„¢", "Vermont Ale (WLP4000)", "Wyeast 3638 Bavarian Wheat"]
Base_Mal_Count = [1, 3]
Base_Mal_Range = [3.5, 7.5]
Mal_Range = [0.1, 1.0]
Mal_Count = [2, 6]
Hop_Range = [10.0, 400.0]
Hop_Count = [1, 6]
Temperature_Range = [63, 69]
Mash_Time = [65, 75, 90]
Fermentation_Range = [19, 21]

# randomizing the number of ingredients
for row in range(2, Beer_Count+2):
    mal_count = random.randint(Mal_Count[0], Mal_Count[1])
    base_mal = random.choice(Base_Mals)
    hop_count = random.randint(Hop_Count[0], Hop_Count[1])
    yeast = random.choice(Yeasts)
    time = random.choice(Mash_Time)
    tempr = random.randint(Temperature_Range[0], Temperature_Range[1])
    ferm = random.randint(Fermentation_Range[0], Fermentation_Range[1])
    col = 1
# Setting up the "r" variable, to decide if we should generate a value to current cell
    while col <= len(names):
        col_value = get_column_letter(col)
        cell_value = ws[get_column_letter(col)+str(row)].value
        r = random.randint(0, 5)
        if cell_value == 0:
            item = 0
        else:
            item = cell_value
        # Mals
        if col < 43:
            if ws[col_value+"1"].value == base_mal:
                item = round(random.uniform(Base_Mal_Range[0], Base_Mal_Range[1]), 1)
                mal_count -= 1
            elif ws[col_value+"1"].value != base_mal and r == 5:
                item = round(random.uniform(Mal_Range[0], Mal_Range[1]), 1)
                mal_count -= 1

        # Hops
        if col >= 43 and col < 110 and hop_count > 0 and r == 1:
            item = round(random.uniform(Hop_Range[0], Hop_Range[1]), 1)
            hop_count -= 1
        # Yeasts
        if col >= 110 and col < 120 and ws[get_column_letter(col)+"1"].value == yeast:
            item = 1
        # Tempr
        if col == 120:
            item = tempr
        # Time
        if col == 121:
            item = time
        # Ferment
        if col == 122:
            item = ferm

        ws[get_column_letter(col)+str(row)].value = item

        # Checking if our Conditions, if we didn't generate enough, than start over from current ingredient type
        if col == 42 and mal_count > 0:
            col = 1
        if col == 119 and hop_count > 0:
            col = 43
        if mal_count == 0 and col < 43:
            col = 43
        if hop_count == 0 and col < 110:
            col = 110
        else:
            col += 1

# Setting up boundaries to the values

for row in range(2, Beer_Count+2):
    malsum = []
    hopsum = []

#Checking every cells value and adding them to malsum or hopsum
    for col in range(1,43):
        cell_value = ws[get_column_letter(col) + str(row)].value
        if ws[get_column_letter(col) + str(row)].value != 0 and (type(ws[get_column_letter(col) + str(row)].value) is int or type(ws[get_column_letter(col) + str(row)].value) is float):
            malsum.append(ws[get_column_letter(col) + str(row)].value)
    for col in range(43,110):
        if ws[get_column_letter(col) + str(row)].value != 0 and (type(ws[get_column_letter(col) + str(row)].value) is int or type(ws[get_column_letter(col) + str(row)].value) is float):
            hopsum.append(ws[get_column_letter(col) + str(row)].value)
    col=1
    while col <= len(names):
        if col<43:
            if sum(malsum)<8: # adding random values to malsum's values if it's too low
                col=1
                for i in range(1,43):
                    list_value = ws[get_column_letter(i) + str(row)].value
                    tp = 0
                    for j in range(len(malsum)):
                        randmal = round(random.uniform(Mal_Range[0], Mal_Range[1]), 1)
                        if list_value == malsum[j] and tp == 0:
                            if ws[get_column_letter(i)+"1"].value in Base_Mals and list_value !=0 and list_value < Base_Mal_Range[1]-randmal:
                                ws[get_column_letter(i) + str(row)].value += randmal
                                malsum[j] = ws[get_column_letter(i) + str(row)].value
                            elif ws[get_column_letter(i)+"1"].value not in Base_Mals and list_value < Mal_Range[1]-randmal:
                                ws[get_column_letter(i) + str(row)].value += randmal
                                malsum[j] = ws[get_column_letter(i) + str(row)].value
                            tp += 1
        if col >= 43 and col<110:
            if sum(hopsum) < 300: # adding random values to hopsum's values if it's too low
                col = 43
                for i in range(43, 110):
                    list_value = ws[get_column_letter(i) + str(row)].value
                    tp = 0
                    for j in range(len(hopsum)):
                        randhop = round(random.uniform(Hop_Range[0], Hop_Range[1]), 1)
                        if list_value == hopsum[j] and tp == 0:
                            if list_value != 0 and list_value < Hop_Range[1] - randhop:
                                ws[get_column_letter(i) + str(row)].value += randhop
                                hopsum[j] = ws[get_column_letter(i) + str(row)].value
                            tp += 1
            if sum(hopsum) > 500: # reducing random values to hopsum's values if it's too high
                col = 43
                for i in range(43, 110):
                    list_value = ws[get_column_letter(i) + str(row)].value
                    tp = 0
                    for j in range(len(hopsum)):
                        randhop = round(random.uniform(Hop_Range[0], Hop_Range[1]), 1)
                        if list_value == hopsum[j] and tp == 0:
                            if list_value != 0 and list_value > Hop_Range[0] + randhop:
                                ws[get_column_letter(i) + str(row)].value -= randhop
                                hopsum[j] = ws[get_column_letter(i) + str(row)].value
                            tp += 1

        col += 1

wb.save(File_Name+'.xlsx')


